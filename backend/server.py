from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, UploadFile, File, Depends, Query
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import httpx
import base64
import io
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import json

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============ MODELS ============

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    profession: Optional[str] = None  # physician, np_pa, nurse
    npi_number: Optional[str] = None  # National Provider Identifier
    npi_verified: bool = False  # Whether NPI has been validated
    npi_data: Optional[Dict[str, Any]] = None  # Data from NPPES registry
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserSession(BaseModel):
    model_config = ConfigDict(extra="ignore")
    session_id: str
    user_id: str
    session_token: str
    expires_at: datetime
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Certificate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    certificate_id: str = Field(default_factory=lambda: f"cert_{uuid.uuid4().hex[:12]}")
    user_id: str
    title: str
    provider: str
    credits: float
    credit_types: List[str] = []  # Multiple credit types e.g., ["ama_cat1", "moc"]
    credit_type: Optional[str] = None  # Legacy field for backwards compatibility
    subject: Optional[str] = None
    completion_date: str
    expiration_date: Optional[str] = None
    certificate_number: Optional[str] = None
    image_url: Optional[str] = None
    ocr_status: str = "none"  # none, processing, completed, failed
    ocr_data: Optional[Dict[str, Any]] = None
    eeds_imported: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CertificateCreate(BaseModel):
    title: str
    provider: str
    credits: float
    credit_types: List[str] = []  # Multiple credit types
    credit_type: Optional[str] = None  # Legacy support
    subject: Optional[str] = None
    completion_date: str
    expiration_date: Optional[str] = None
    certificate_number: Optional[str] = None

class Requirement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    requirement_id: str = Field(default_factory=lambda: f"req_{uuid.uuid4().hex[:12]}")
    user_id: str
    name: str
    requirement_type: str  # license_renewal, board_recert, hospital, personal
    credit_types: List[str] = []  # Multiple credit types that can satisfy this requirement
    credit_type: Optional[str] = None  # Legacy field
    providers: List[str] = []  # Filter by specific providers (e.g., "ACCME", "Hospital XYZ")
    subjects: List[str] = []  # Filter by specific subjects (e.g., "Cardiology", "Pain Management")
    credits_required: float
    credits_earned: float = 0
    start_year: Optional[int] = None  # Year range start
    end_year: Optional[int] = None  # Year range end
    due_date: str
    notes: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RequirementCreate(BaseModel):
    name: str
    requirement_type: str
    credit_types: List[str] = []
    credit_type: Optional[str] = None
    providers: List[str] = []
    subjects: List[str] = []
    credits_required: float
    start_year: Optional[int] = None
    end_year: Optional[int] = None
    due_date: str
    notes: Optional[str] = None

class RequirementUpdate(BaseModel):
    name: Optional[str] = None
    credit_types: Optional[List[str]] = None
    providers: Optional[List[str]] = None
    subjects: Optional[List[str]] = None
    credits_required: Optional[float] = None
    start_year: Optional[int] = None
    end_year: Optional[int] = None
    due_date: Optional[str] = None
    notes: Optional[str] = None
    is_active: Optional[bool] = None

class CustomCreditType(BaseModel):
    model_config = ConfigDict(extra="ignore")
    credit_type_id: str = Field(default_factory=lambda: f"custom_{uuid.uuid4().hex[:8]}")
    user_id: str
    name: str
    description: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ============ NEW MODELS FOR EEDS PARITY ============

# Self-Reported Credits (journal clubs, self-study, presentations, etc.)
class SelfReportedCredit(BaseModel):
    model_config = ConfigDict(extra="ignore")
    credit_id: str = Field(default_factory=lambda: f"self_{uuid.uuid4().hex[:12]}")
    user_id: str
    activity_type: str  # self_study, journal_club, presentation, teaching, manuscript, peer_review, other
    title: str
    description: Optional[str] = None
    credits: float
    credit_types: List[str] = []
    completion_date: str
    hours_spent: Optional[float] = None  # Time spent on activity
    reference_url: Optional[str] = None  # Link to article, resource, etc.
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SelfReportedCreditCreate(BaseModel):
    activity_type: str
    title: str
    description: Optional[str] = None
    credits: float
    credit_types: List[str] = []
    completion_date: str
    hours_spent: Optional[float] = None
    reference_url: Optional[str] = None

# CME Events/Calendar
class CMEEvent(BaseModel):
    model_config = ConfigDict(extra="ignore")
    event_id: str = Field(default_factory=lambda: f"evt_{uuid.uuid4().hex[:12]}")
    user_id: str
    title: str
    description: Optional[str] = None
    provider: str
    location: Optional[str] = None  # Physical location or "Virtual"
    event_url: Optional[str] = None
    start_date: str
    end_date: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    credits_available: Optional[float] = None
    credit_types: List[str] = []
    registration_url: Optional[str] = None
    cost: Optional[float] = None
    is_registered: bool = False
    is_attended: bool = False
    passcode: Optional[str] = None  # 6-digit event passcode for sign-in
    notes: Optional[str] = None
    reminder_sent: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CMEEventCreate(BaseModel):
    title: str
    description: Optional[str] = None
    provider: str
    location: Optional[str] = None
    event_url: Optional[str] = None
    start_date: str
    end_date: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    credits_available: Optional[float] = None
    credit_types: List[str] = []
    registration_url: Optional[str] = None
    cost: Optional[float] = None
    notes: Optional[str] = None

# Event Evaluations
class Evaluation(BaseModel):
    model_config = ConfigDict(extra="ignore")
    evaluation_id: str = Field(default_factory=lambda: f"eval_{uuid.uuid4().hex[:12]}")
    user_id: str
    certificate_id: Optional[str] = None  # Link to certificate if from formal CME
    event_id: Optional[str] = None  # Link to event
    title: str  # Activity title
    overall_rating: int  # 1-5 scale
    content_quality: Optional[int] = None  # 1-5
    speaker_effectiveness: Optional[int] = None  # 1-5
    relevance_to_practice: Optional[int] = None  # 1-5
    would_recommend: Optional[bool] = None
    learning_objectives_met: Optional[bool] = None
    comments: Optional[str] = None
    improvement_suggestions: Optional[str] = None
    practice_change_planned: Optional[str] = None  # What will you do differently?
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EvaluationCreate(BaseModel):
    certificate_id: Optional[str] = None
    event_id: Optional[str] = None
    title: str
    overall_rating: int
    content_quality: Optional[int] = None
    speaker_effectiveness: Optional[int] = None
    relevance_to_practice: Optional[int] = None
    would_recommend: Optional[bool] = None
    learning_objectives_met: Optional[bool] = None
    comments: Optional[str] = None
    improvement_suggestions: Optional[str] = None
    practice_change_planned: Optional[str] = None

# Speaker Disclosures (for tracking COI)
class SpeakerDisclosure(BaseModel):
    model_config = ConfigDict(extra="ignore")
    disclosure_id: str = Field(default_factory=lambda: f"disc_{uuid.uuid4().hex[:12]}")
    user_id: str
    event_id: Optional[str] = None
    certificate_id: Optional[str] = None
    speaker_name: str
    speaker_credentials: Optional[str] = None  # MD, PhD, etc.
    has_conflicts: bool = False
    disclosure_text: Optional[str] = None  # Full disclosure statement
    financial_relationships: List[Dict[str, Any]] = []  # [{company, relationship_type, amount}]
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SpeakerDisclosureCreate(BaseModel):
    event_id: Optional[str] = None
    certificate_id: Optional[str] = None
    speaker_name: str
    speaker_credentials: Optional[str] = None
    has_conflicts: bool = False
    disclosure_text: Optional[str] = None
    financial_relationships: List[Dict[str, Any]] = []

# Course Materials (attachments)
class CourseMaterial(BaseModel):
    model_config = ConfigDict(extra="ignore")
    material_id: str = Field(default_factory=lambda: f"mat_{uuid.uuid4().hex[:12]}")
    user_id: str
    certificate_id: Optional[str] = None
    event_id: Optional[str] = None
    title: str
    material_type: str  # handout, slides, video, article, other
    file_url: Optional[str] = None  # Base64 or external URL
    file_name: Optional[str] = None
    file_size: Optional[int] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Self-Reported Activity Types
SELF_REPORTED_TYPES = [
    {"id": "self_study", "name": "Self-Study", "description": "Independent learning from books, articles, online resources"},
    {"id": "journal_club", "name": "Journal Club", "description": "Participation in journal club discussions"},
    {"id": "presentation", "name": "Presentation/Lecture", "description": "Giving a presentation or lecture"},
    {"id": "teaching", "name": "Teaching", "description": "Teaching medical students, residents, or peers"},
    {"id": "manuscript", "name": "Manuscript Review/Writing", "description": "Reviewing or writing medical manuscripts"},
    {"id": "peer_review", "name": "Peer Review", "description": "Peer review of articles or case presentations"},
    {"id": "quality_improvement", "name": "Quality Improvement", "description": "QI projects and activities"},
    {"id": "tumor_board", "name": "Tumor Board", "description": "Tumor board participation"},
    {"id": "grand_rounds", "name": "Grand Rounds", "description": "Attending grand rounds presentations"},
    {"id": "case_conference", "name": "Case Conference", "description": "Case discussion and review"},
    {"id": "other", "name": "Other", "description": "Other educational activities"}
]

# CME Types by Profession
CME_TYPES = {
    "physician": [
        {"id": "ama_cat1", "name": "AMA PRA Category 1", "description": "Gold standard for physician CME"},
        {"id": "ama_cat2", "name": "AMA PRA Category 2", "description": "Self-reported educational activities"},
        {"id": "aoa_1a", "name": "AOA Category 1-A", "description": "Osteopathic medical teaching"},
        {"id": "aoa_1b", "name": "AOA Category 1-B", "description": "Board certification activities"},
        {"id": "moc", "name": "MOC/MOL", "description": "Maintenance of Certification/Licensure"},
        {"id": "self_assessment", "name": "Self-Assessment", "description": "Knowledge self-assessment"},
        {"id": "ethics", "name": "Medical Ethics", "description": "Ethics credits"},
        {"id": "pain_mgmt", "name": "Pain Management", "description": "Pain management/opioid prescribing"},
    ],
    "np_pa": [
        {"id": "aanp_contact", "name": "AANP Contact Hours", "description": "NP contact hours"},
        {"id": "aapa_cat1", "name": "AAPA Category 1", "description": "PA Category 1 credits"},
        {"id": "ama_cat1", "name": "AMA PRA Category 1", "description": "Accepted for NP/PA"},
        {"id": "pharmacology", "name": "Pharmacology CE", "description": "Pharmacology continuing education"},
        {"id": "ancc_contact", "name": "ANCC Contact Hours", "description": "Nursing contact hours"},
        {"id": "self_assessment", "name": "Self-Assessment", "description": "Knowledge self-assessment"},
    ],
    "nurse": [
        {"id": "ancc_contact", "name": "ANCC Contact Hours", "description": "Primary nursing CE"},
        {"id": "cne", "name": "CNE Credits", "description": "Continuing nursing education"},
        {"id": "pharmacology", "name": "Pharmacology CE", "description": "Required for NPs/CNSs"},
        {"id": "specialty", "name": "Specialty CE", "description": "Specialty-specific education"},
        {"id": "ethics", "name": "Nursing Ethics", "description": "Ethics credits"},
        {"id": "cultural", "name": "Cultural Competency", "description": "Cultural competency training"},
    ]
}

# ============ AUTH HELPERS ============

async def get_current_user(request: Request) -> User:
    """Get current user from session token"""
    session_token = request.cookies.get("session_token")
    if not session_token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            session_token = auth_header.split(" ")[1]
    
    if not session_token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    expires_at = session["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")
    
    user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    return User(**user)

# ============ AUTH ROUTES ============

@api_router.post("/auth/session")
async def exchange_session(request: Request, response: Response):
    """Exchange session_id for session_token"""
    body = await request.json()
    session_id = body.get("session_id")
    
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    
    # Call Emergent Auth to get user data
    async with httpx.AsyncClient() as client:
        auth_response = await client.get(
            "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
            headers={"X-Session-ID": session_id}
        )
    
    if auth_response.status_code != 200:
        raise HTTPException(status_code=401, detail="Invalid session_id")
    
    auth_data = auth_response.json()
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    session_token = auth_data.get("session_token", f"st_{uuid.uuid4().hex}")
    
    # Check if user exists
    existing_user = await db.users.find_one({"email": auth_data["email"]}, {"_id": 0})
    
    if existing_user:
        user_id = existing_user["user_id"]
        # Update user data
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {
                "name": auth_data["name"],
                "picture": auth_data.get("picture"),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    else:
        # Create new user
        new_user = {
            "user_id": user_id,
            "email": auth_data["email"],
            "name": auth_data["name"],
            "picture": auth_data.get("picture"),
            "profession": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(new_user)
    
    # Create session
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    session_doc = {
        "session_id": str(uuid.uuid4()),
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Remove old sessions for this user
    await db.user_sessions.delete_many({"user_id": user_id})
    await db.user_sessions.insert_one(session_doc)
    
    # Set cookie
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=7 * 24 * 60 * 60,
        path="/"
    )
    
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    return user

@api_router.get("/auth/me")
async def get_me(user: User = Depends(get_current_user)):
    """Get current authenticated user"""
    return user.model_dump()

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    """Logout user"""
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_many({"session_token": session_token})
    
    response.delete_cookie(key="session_token", path="/", secure=True, samesite="none")
    return {"message": "Logged out successfully"}

# ============ USER ROUTES ============

@api_router.put("/users/profession")
async def update_profession(request: Request, user: User = Depends(get_current_user)):
    """Update user profession"""
    body = await request.json()
    profession = body.get("profession")
    
    if profession not in ["physician", "np_pa", "nurse"]:
        raise HTTPException(status_code=400, detail="Invalid profession")
    
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$set": {"profession": profession, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    updated_user = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    return updated_user

@api_router.get("/users/profile")
async def get_profile(user: User = Depends(get_current_user)):
    """Get user profile with stats"""
    # Get certificate stats
    total_certs = await db.certificates.count_documents({"user_id": user.user_id})
    
    # Get total credits
    pipeline = [
        {"$match": {"user_id": user.user_id}},
        {"$group": {"_id": None, "total_credits": {"$sum": "$credits"}}}
    ]
    result = await db.certificates.aggregate(pipeline).to_list(1)
    total_credits = result[0]["total_credits"] if result else 0
    
    # Get active requirements
    active_reqs = await db.requirements.count_documents({"user_id": user.user_id, "is_active": True})
    
    return {
        **user.model_dump(),
        "stats": {
            "total_certificates": total_certs,
            "total_credits": total_credits,
            "active_requirements": active_reqs
        }
    }

# ============ NPI VALIDATION ============

def validate_npi_checksum(npi: str) -> bool:
    """Validate NPI using Luhn algorithm (ISO/IEC 7812)"""
    if not npi or len(npi) != 10 or not npi.isdigit():
        return False
    
    # Prefix with 80840 for healthcare NPI
    prefixed = "80840" + npi
    
    # Luhn algorithm
    total = 0
    for i, digit in enumerate(reversed(prefixed)):
        d = int(digit)
        if i % 2 == 1:  # Double every second digit from right
            d *= 2
            if d > 9:
                d -= 9
        total += d
    
    return total % 10 == 0

async def lookup_npi_registry(npi: str) -> Optional[Dict[str, Any]]:
    """Look up NPI in NPPES registry"""
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(
                f"https://npiregistry.cms.hhs.gov/api/?number={npi}&version=2.1",
                timeout=10.0
            )
            
            if response.status_code == 200:
                data = response.json()
                if data.get("result_count", 0) > 0:
                    result = data["results"][0]
                    basic = result.get("basic", {})
                    
                    # Extract relevant info
                    return {
                        "npi": result.get("number"),
                        "entity_type": result.get("enumeration_type"),
                        "first_name": basic.get("first_name"),
                        "last_name": basic.get("last_name"),
                        "organization_name": basic.get("organization_name"),
                        "credential": basic.get("credential"),
                        "status": basic.get("status"),
                        "enumeration_date": basic.get("enumeration_date"),
                        "last_updated": basic.get("last_updated"),
                        "taxonomies": [
                            {
                                "code": t.get("code"),
                                "desc": t.get("desc"),
                                "primary": t.get("primary")
                            }
                            for t in result.get("taxonomies", [])
                        ]
                    }
        return None
    except Exception as e:
        logger.error(f"NPI lookup error: {e}")
        return None

@api_router.post("/users/npi/validate")
async def validate_npi(request: Request, user: User = Depends(get_current_user)):
    """Validate and link NPI number to profile"""
    body = await request.json()
    npi = body.get("npi", "").strip()
    
    if not npi:
        raise HTTPException(status_code=400, detail="NPI number is required")
    
    # Format validation
    if not validate_npi_checksum(npi):
        raise HTTPException(status_code=400, detail="Invalid NPI format. Must be a valid 10-digit NPI number.")
    
    # Lookup in NPPES registry
    npi_data = await lookup_npi_registry(npi)
    
    if not npi_data:
        raise HTTPException(status_code=404, detail="NPI not found in NPPES registry. Please verify the number.")
    
    # Update user with NPI data
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$set": {
            "npi_number": npi,
            "npi_verified": True,
            "npi_data": npi_data,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated_user = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    return {
        "message": "NPI verified successfully",
        "user": updated_user,
        "npi_data": npi_data
    }

@api_router.delete("/users/npi")
async def remove_npi(user: User = Depends(get_current_user)):
    """Remove NPI number from profile"""
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$set": {
            "npi_number": None,
            "npi_verified": False,
            "npi_data": None,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated_user = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    return {"message": "NPI removed", "user": updated_user}

# ============ CME TYPES ROUTES ============

@api_router.get("/cme-types")
async def get_cme_types(user: User = Depends(get_current_user)):
    """Get CME types for user's profession including custom types"""
    profession = user.profession or "physician"
    standard_types = CME_TYPES.get(profession, CME_TYPES["physician"])
    
    # Get user's custom credit types
    custom_types = await db.custom_credit_types.find(
        {"user_id": user.user_id},
        {"_id": 0}
    ).to_list(100)
    
    # Convert custom types to same format
    custom_formatted = [
        {"id": ct["credit_type_id"], "name": ct["name"], "description": ct.get("description", "Custom credit type"), "is_custom": True}
        for ct in custom_types
    ]
    
    return standard_types + custom_formatted

@api_router.get("/cme-types/all")
async def get_all_cme_types():
    """Get all CME types"""
    return CME_TYPES

@api_router.post("/cme-types/custom")
async def create_custom_credit_type(request: Request, user: User = Depends(get_current_user)):
    """Create a custom credit type"""
    body = await request.json()
    name = body.get("name", "").strip()
    description = body.get("description", "")
    
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")
    
    # Check for duplicate name
    existing = await db.custom_credit_types.find_one(
        {"user_id": user.user_id, "name": {"$regex": f"^{name}$", "$options": "i"}},
        {"_id": 0}
    )
    if existing:
        raise HTTPException(status_code=400, detail="A custom credit type with this name already exists")
    
    custom_type = CustomCreditType(
        user_id=user.user_id,
        name=name,
        description=description
    )
    
    type_dict = custom_type.model_dump()
    type_dict["created_at"] = type_dict["created_at"].isoformat()
    
    await db.custom_credit_types.insert_one(type_dict)
    type_dict.pop("_id", None)
    
    return type_dict

@api_router.get("/cme-types/custom")
async def get_custom_credit_types(user: User = Depends(get_current_user)):
    """Get user's custom credit types"""
    custom_types = await db.custom_credit_types.find(
        {"user_id": user.user_id},
        {"_id": 0}
    ).to_list(100)
    return custom_types

@api_router.delete("/cme-types/custom/{credit_type_id}")
async def delete_custom_credit_type(credit_type_id: str, user: User = Depends(get_current_user)):
    """Delete a custom credit type"""
    result = await db.custom_credit_types.delete_one(
        {"credit_type_id": credit_type_id, "user_id": user.user_id}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Custom credit type not found")
    return {"message": "Custom credit type deleted"}

# ============ SELF-REPORTED CREDITS ROUTES ============

@api_router.get("/self-reported-types")
async def get_self_reported_types():
    """Get available self-reported activity types"""
    return SELF_REPORTED_TYPES

@api_router.get("/self-reported")
async def get_self_reported_credits(
    user: User = Depends(get_current_user),
    year: Optional[int] = None
):
    """Get user's self-reported credits"""
    query = {"user_id": user.user_id}
    if year:
        query["completion_date"] = {"$regex": f"^{year}"}
    
    credits = await db.self_reported_credits.find(query, {"_id": 0}).sort("completion_date", -1).to_list(500)
    return credits

@api_router.post("/self-reported")
async def create_self_reported_credit(data: SelfReportedCreditCreate, user: User = Depends(get_current_user)):
    """Create a self-reported credit entry"""
    credit = SelfReportedCredit(user_id=user.user_id, **data.model_dump())
    
    credit_dict = credit.model_dump()
    credit_dict["created_at"] = credit_dict["created_at"].isoformat()
    credit_dict["updated_at"] = credit_dict["updated_at"].isoformat()
    
    await db.self_reported_credits.insert_one(credit_dict)
    credit_dict.pop("_id", None)
    
    # Update requirement progress (self-reported credits count too)
    await update_requirement_progress(user.user_id)
    
    return credit_dict

@api_router.put("/self-reported/{credit_id}")
async def update_self_reported_credit(credit_id: str, request: Request, user: User = Depends(get_current_user)):
    """Update a self-reported credit"""
    body = await request.json()
    
    result = await db.self_reported_credits.update_one(
        {"credit_id": credit_id, "user_id": user.user_id},
        {"$set": {**body, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Self-reported credit not found")
    
    await update_requirement_progress(user.user_id)
    
    credit = await db.self_reported_credits.find_one({"credit_id": credit_id}, {"_id": 0})
    return credit

@api_router.delete("/self-reported/{credit_id}")
async def delete_self_reported_credit(credit_id: str, user: User = Depends(get_current_user)):
    """Delete a self-reported credit"""
    result = await db.self_reported_credits.delete_one(
        {"credit_id": credit_id, "user_id": user.user_id}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Self-reported credit not found")
    
    await update_requirement_progress(user.user_id)
    return {"message": "Self-reported credit deleted"}

# ============ CME EVENTS/CALENDAR ROUTES ============

@api_router.get("/events")
async def get_events(
    user: User = Depends(get_current_user),
    upcoming: bool = False,
    past: bool = False
):
    """Get user's CME events"""
    query = {"user_id": user.user_id}
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    if upcoming:
        query["start_date"] = {"$gte": today}
    elif past:
        query["start_date"] = {"$lt": today}
    
    events = await db.cme_events.find(query, {"_id": 0}).sort("start_date", 1).to_list(500)
    return events

@api_router.post("/events")
async def create_event(data: CMEEventCreate, user: User = Depends(get_current_user)):
    """Create a CME event"""
    # Generate a 6-digit passcode for sign-in
    import random
    passcode = str(random.randint(100000, 999999))
    
    event = CMEEvent(user_id=user.user_id, passcode=passcode, **data.model_dump())
    
    event_dict = event.model_dump()
    event_dict["created_at"] = event_dict["created_at"].isoformat()
    event_dict["updated_at"] = event_dict["updated_at"].isoformat()
    
    await db.cme_events.insert_one(event_dict)
    event_dict.pop("_id", None)
    
    return event_dict

@api_router.get("/events/{event_id}")
async def get_event(event_id: str, user: User = Depends(get_current_user)):
    """Get a specific event"""
    event = await db.cme_events.find_one(
        {"event_id": event_id, "user_id": user.user_id},
        {"_id": 0}
    )
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    return event

@api_router.put("/events/{event_id}")
async def update_event(event_id: str, request: Request, user: User = Depends(get_current_user)):
    """Update an event"""
    body = await request.json()
    
    result = await db.cme_events.update_one(
        {"event_id": event_id, "user_id": user.user_id},
        {"$set": {**body, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    
    event = await db.cme_events.find_one({"event_id": event_id}, {"_id": 0})
    return event

@api_router.delete("/events/{event_id}")
async def delete_event(event_id: str, user: User = Depends(get_current_user)):
    """Delete an event"""
    result = await db.cme_events.delete_one(
        {"event_id": event_id, "user_id": user.user_id}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    
    return {"message": "Event deleted"}

@api_router.post("/events/{event_id}/register")
async def toggle_event_registration(event_id: str, user: User = Depends(get_current_user)):
    """Toggle event registration status"""
    event = await db.cme_events.find_one({"event_id": event_id, "user_id": user.user_id})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    new_status = not event.get("is_registered", False)
    await db.cme_events.update_one(
        {"event_id": event_id},
        {"$set": {"is_registered": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"is_registered": new_status}

@api_router.post("/events/{event_id}/attend")
async def mark_event_attended(event_id: str, request: Request, user: User = Depends(get_current_user)):
    """Mark event as attended (with optional passcode verification)"""
    body = await request.json()
    passcode = body.get("passcode")
    
    event = await db.cme_events.find_one({"event_id": event_id, "user_id": user.user_id})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    # If event has a passcode, verify it
    if event.get("passcode") and passcode:
        if event["passcode"] != passcode:
            raise HTTPException(status_code=400, detail="Invalid passcode")
    
    await db.cme_events.update_one(
        {"event_id": event_id},
        {"$set": {"is_attended": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Event marked as attended", "is_attended": True}

@api_router.post("/events/sign-in")
async def sign_in_to_event(request: Request, user: User = Depends(get_current_user)):
    """Sign in to an event using passcode"""
    body = await request.json()
    passcode = body.get("passcode", "").strip()
    
    if not passcode or len(passcode) != 6:
        raise HTTPException(status_code=400, detail="Please enter a valid 6-digit passcode")
    
    # Find event with this passcode
    event = await db.cme_events.find_one(
        {"user_id": user.user_id, "passcode": passcode},
        {"_id": 0}
    )
    
    if not event:
        raise HTTPException(status_code=404, detail="No event found with this passcode")
    
    # Mark as attended
    await db.cme_events.update_one(
        {"event_id": event["event_id"]},
        {"$set": {"is_attended": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": f"Successfully signed in to: {event['title']}", "event": event}

# ============ EVALUATIONS ROUTES ============

@api_router.get("/evaluations")
async def get_evaluations(user: User = Depends(get_current_user)):
    """Get user's evaluations"""
    evaluations = await db.evaluations.find(
        {"user_id": user.user_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    return evaluations

@api_router.post("/evaluations")
async def create_evaluation(data: EvaluationCreate, user: User = Depends(get_current_user)):
    """Create an evaluation"""
    evaluation = Evaluation(user_id=user.user_id, **data.model_dump())
    
    eval_dict = evaluation.model_dump()
    eval_dict["created_at"] = eval_dict["created_at"].isoformat()
    
    await db.evaluations.insert_one(eval_dict)
    eval_dict.pop("_id", None)
    
    return eval_dict

@api_router.get("/evaluations/{evaluation_id}")
async def get_evaluation(evaluation_id: str, user: User = Depends(get_current_user)):
    """Get a specific evaluation"""
    evaluation = await db.evaluations.find_one(
        {"evaluation_id": evaluation_id, "user_id": user.user_id},
        {"_id": 0}
    )
    if not evaluation:
        raise HTTPException(status_code=404, detail="Evaluation not found")
    return evaluation

@api_router.delete("/evaluations/{evaluation_id}")
async def delete_evaluation(evaluation_id: str, user: User = Depends(get_current_user)):
    """Delete an evaluation"""
    result = await db.evaluations.delete_one(
        {"evaluation_id": evaluation_id, "user_id": user.user_id}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Evaluation not found")
    
    return {"message": "Evaluation deleted"}

# ============ SPEAKER DISCLOSURES ROUTES ============

@api_router.get("/disclosures")
async def get_disclosures(user: User = Depends(get_current_user)):
    """Get speaker disclosures"""
    disclosures = await db.speaker_disclosures.find(
        {"user_id": user.user_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    return disclosures

@api_router.post("/disclosures")
async def create_disclosure(data: SpeakerDisclosureCreate, user: User = Depends(get_current_user)):
    """Create a speaker disclosure"""
    disclosure = SpeakerDisclosure(user_id=user.user_id, **data.model_dump())
    
    disc_dict = disclosure.model_dump()
    disc_dict["created_at"] = disc_dict["created_at"].isoformat()
    
    await db.speaker_disclosures.insert_one(disc_dict)
    disc_dict.pop("_id", None)
    
    return disc_dict

@api_router.delete("/disclosures/{disclosure_id}")
async def delete_disclosure(disclosure_id: str, user: User = Depends(get_current_user)):
    """Delete a speaker disclosure"""
    result = await db.speaker_disclosures.delete_one(
        {"disclosure_id": disclosure_id, "user_id": user.user_id}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Disclosure not found")
    
    return {"message": "Disclosure deleted"}

# ============ COURSE MATERIALS ROUTES ============

@api_router.get("/materials")
async def get_materials(
    user: User = Depends(get_current_user),
    certificate_id: Optional[str] = None,
    event_id: Optional[str] = None
):
    """Get course materials"""
    query = {"user_id": user.user_id}
    if certificate_id:
        query["certificate_id"] = certificate_id
    if event_id:
        query["event_id"] = event_id
    
    materials = await db.course_materials.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return materials

@api_router.post("/materials")
async def upload_material(
    certificate_id: Optional[str] = None,
    event_id: Optional[str] = None,
    title: str = "",
    material_type: str = "other",
    notes: str = "",
    file: UploadFile = File(...),
    user: User = Depends(get_current_user)
):
    """Upload a course material"""
    content = await file.read()
    base64_content = base64.b64encode(content).decode('utf-8')
    
    material = CourseMaterial(
        user_id=user.user_id,
        certificate_id=certificate_id,
        event_id=event_id,
        title=title or file.filename,
        material_type=material_type,
        file_url=f"data:{file.content_type};base64,{base64_content}",
        file_name=file.filename,
        file_size=len(content),
        notes=notes
    )
    
    mat_dict = material.model_dump()
    mat_dict["created_at"] = mat_dict["created_at"].isoformat()
    
    await db.course_materials.insert_one(mat_dict)
    mat_dict.pop("_id", None)
    
    return mat_dict

@api_router.delete("/materials/{material_id}")
async def delete_material(material_id: str, user: User = Depends(get_current_user)):
    """Delete a course material"""
    result = await db.course_materials.delete_one(
        {"material_id": material_id, "user_id": user.user_id}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Material not found")
    
    return {"message": "Material deleted"}

# ============ CERTIFICATE ROUTES ============

@api_router.get("/certificates")
async def get_certificates(
    user: User = Depends(get_current_user),
    credit_type: Optional[str] = None,
    year: Optional[int] = None
):
    """Get user's certificates"""
    query = {"user_id": user.user_id}
    
    if credit_type:
        # Search in both credit_types array and legacy credit_type field
        query["$or"] = [
            {"credit_types": credit_type},
            {"credit_type": credit_type}
        ]
    
    if year:
        query["completion_date"] = {"$regex": f"^{year}"}
    
    certificates = await db.certificates.find(query, {"_id": 0}).sort("completion_date", -1).to_list(1000)
    
    # Normalize credit_types for backwards compatibility
    for cert in certificates:
        if not cert.get("credit_types") and cert.get("credit_type"):
            cert["credit_types"] = [cert["credit_type"]]
    
    return certificates

@api_router.post("/certificates")
async def create_certificate(cert_data: CertificateCreate, user: User = Depends(get_current_user)):
    """Create a new certificate"""
    data = cert_data.model_dump()
    
    # Handle credit_types - ensure it's populated
    if not data.get("credit_types") and data.get("credit_type"):
        data["credit_types"] = [data["credit_type"]]
    elif data.get("credit_types") and not data.get("credit_type"):
        data["credit_type"] = data["credit_types"][0] if data["credit_types"] else None
    
    cert = Certificate(user_id=user.user_id, **data)
    
    cert_dict = cert.model_dump()
    cert_dict["created_at"] = cert_dict["created_at"].isoformat()
    cert_dict["updated_at"] = cert_dict["updated_at"].isoformat()
    
    await db.certificates.insert_one(cert_dict)
    cert_dict.pop("_id", None)  # Remove MongoDB's _id to avoid serialization error
    
    # Update requirement progress
    await update_requirement_progress(user.user_id)
    
    return cert_dict

@api_router.get("/certificates/{certificate_id}")
async def get_certificate(certificate_id: str, user: User = Depends(get_current_user)):
    """Get a specific certificate"""
    cert = await db.certificates.find_one(
        {"certificate_id": certificate_id, "user_id": user.user_id},
        {"_id": 0}
    )
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")
    return cert

@api_router.put("/certificates/{certificate_id}")
async def update_certificate(certificate_id: str, request: Request, user: User = Depends(get_current_user)):
    """Update a certificate"""
    body = await request.json()
    
    result = await db.certificates.update_one(
        {"certificate_id": certificate_id, "user_id": user.user_id},
        {"$set": {**body, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Certificate not found")
    
    # Update requirement progress
    await update_requirement_progress(user.user_id)
    
    cert = await db.certificates.find_one({"certificate_id": certificate_id}, {"_id": 0})
    return cert

@api_router.delete("/certificates/{certificate_id}")
async def delete_certificate(certificate_id: str, user: User = Depends(get_current_user)):
    """Delete a certificate"""
    result = await db.certificates.delete_one(
        {"certificate_id": certificate_id, "user_id": user.user_id}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Certificate not found")
    
    # Update requirement progress
    await update_requirement_progress(user.user_id)
    
    return {"message": "Certificate deleted"}

@api_router.post("/certificates/upload")
async def upload_certificate(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user)
):
    """Upload certificate image/PDF for OCR processing"""
    # Read file content
    content = await file.read()
    base64_content = base64.b64encode(content).decode('utf-8')
    
    # Create certificate with pending OCR
    cert = Certificate(
        user_id=user.user_id,
        title="Processing...",
        provider="Processing...",
        credits=0,
        credit_type="unknown",
        completion_date=datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        ocr_status="processing",
        image_url=f"data:{file.content_type};base64,{base64_content}"
    )
    
    cert_dict = cert.model_dump()
    cert_dict["created_at"] = cert_dict["created_at"].isoformat()
    cert_dict["updated_at"] = cert_dict["updated_at"].isoformat()
    
    await db.certificates.insert_one(cert_dict)
    cert_dict.pop("_id", None)  # Remove MongoDB's _id to avoid serialization error
    
    # Process OCR in background (we'll return immediately and process async)
    # For now, we'll do it synchronously for simplicity
    try:
        ocr_result = await process_certificate_ocr(cert.certificate_id, base64_content, file.content_type)
        return ocr_result
    except Exception as e:
        logger.error(f"OCR processing failed: {e}")
        await db.certificates.update_one(
            {"certificate_id": cert.certificate_id},
            {"$set": {"ocr_status": "failed"}}
        )
        return cert_dict

async def process_certificate_ocr(certificate_id: str, base64_content: str, mime_type: str):
    """Process certificate with GPT-4o vision - enhanced with better error handling and prompting"""
    ocr_error_message = None
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
        
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            ocr_error_message = "OCR service not configured. Please enter details manually."
            raise Exception("EMERGENT_LLM_KEY not configured")
        
        # Handle PDF files by converting to image
        image_base64 = base64_content
        final_mime_type = mime_type
        
        if mime_type == "application/pdf":
            try:
                from pdf2image import convert_from_bytes
                from PIL import Image
                import io
                
                logger.info(f"Converting PDF to image for {certificate_id}")
                
                # Decode the base64 PDF content
                pdf_bytes = base64.b64decode(base64_content)
                
                # Convert PDF to images (first page only for certificates)
                images = convert_from_bytes(pdf_bytes, dpi=150, first_page=1, last_page=1)
                
                if images:
                    # Convert the first page to PNG
                    img_buffer = io.BytesIO()
                    images[0].save(img_buffer, format='PNG', optimize=True)
                    img_buffer.seek(0)
                    
                    # Re-encode as base64
                    image_base64 = base64.b64encode(img_buffer.getvalue()).decode('utf-8')
                    final_mime_type = "image/png"
                    logger.info(f"PDF converted to PNG successfully for {certificate_id}")
                else:
                    ocr_error_message = "Failed to convert PDF to image. Please upload as PNG or JPEG."
                    raise Exception("PDF conversion produced no images")
                    
            except ImportError as e:
                logger.error(f"pdf2image not installed: {e}")
                ocr_error_message = "PDF processing not available. Please upload as PNG or JPEG."
                raise Exception("PDF processing library not available")
            except Exception as e:
                logger.error(f"PDF conversion error for {certificate_id}: {e}")
                ocr_error_message = f"Failed to process PDF: {str(e)}. Please upload as PNG or JPEG."
                raise
        
        # Validate image format for GPT-4o
        supported_formats = ["image/png", "image/jpeg", "image/gif", "image/webp"]
        if final_mime_type not in supported_formats:
            ocr_error_message = f"Unsupported image format: {final_mime_type}. Please upload PNG, JPEG, GIF, or WebP."
            raise Exception(f"Unsupported format: {final_mime_type}")
        
        # Enhanced system prompt for better extraction
        system_prompt = """You are an expert at extracting information from medical CME (Continuing Medical Education) certificates. 

Analyze the certificate image carefully and extract ALL available information. CME certificates typically contain:
- Activity/course title or name
- Provider/sponsor organization (ACCME-accredited organizations, medical associations, hospitals)
- Number of credits earned (look for numbers followed by "credits", "hours", "CME", "AMA PRA Category 1", etc.)
- Credit type (AMA PRA Category 1, Category 2, AANP, AAPA, ANCC, MOC, etc.)
- Completion date or date awarded
- Certificate/reference number
- Medical subject or specialty

IMPORTANT INSTRUCTIONS:
1. Return ONLY a valid JSON object - no explanation, no markdown
2. Use null for any field you cannot determine with confidence
3. For dates, use YYYY-MM-DD format (e.g., "2024-03-15")
4. For credits, extract the numeric value only (e.g., 1.5, not "1.5 credits")
5. For credit_type, use the exact wording from the certificate

JSON format:
{"title": "string or null", "provider": "string or null", "credits": number or null, "credit_type": "string or null", "completion_date": "YYYY-MM-DD or null", "certificate_number": "string or null", "subject": "string or null"}"""

        chat = LlmChat(
            api_key=api_key,
            session_id=f"ocr_{certificate_id}",
            system_message=system_prompt
        ).with_model("openai", "gpt-4o")
        
        # Use the converted image (either original or PDF->PNG)
        image_content = ImageContent(image_base64=image_base64)
        
        response = await chat.send_message(UserMessage(
            text="Extract all CME certificate information from this image. Return only the JSON object.",
            file_contents=[image_content]
        ))
        
        logger.info(f"OCR Response for {certificate_id}: {response[:500]}...")
        
        # Parse response with improved handling
        ocr_data = {}
        parse_error = None
        try:
            # Clean up response - handle various formats
            response_text = response.strip()
            
            # Remove markdown code blocks
            if response_text.startswith("```json"):
                response_text = response_text[7:]
            elif response_text.startswith("```"):
                response_text = response_text[3:]
            if response_text.endswith("```"):
                response_text = response_text[:-3]
            response_text = response_text.strip()
            
            # Try to find JSON in the response if not pure JSON
            if not response_text.startswith("{"):
                import re
                json_match = re.search(r'\{[^{}]*\}', response_text, re.DOTALL)
                if json_match:
                    response_text = json_match.group()
            
            ocr_data = json.loads(response_text)
            
            # Validate we got at least some data
            if not any(ocr_data.get(k) for k in ["title", "provider", "credits"]):
                parse_error = "Could not extract key information"
                
        except json.JSONDecodeError as e:
            parse_error = f"Failed to parse OCR response: {str(e)}"
            logger.warning(f"JSON parse error for {certificate_id}: {e}. Response: {response_text[:200]}")
            ocr_data = {"raw_text": response, "parse_error": str(e)}
        
        # Enhanced credit type mapping with more variations
        credit_type_map = {
            # AMA
            "ama pra category 1": "ama_cat1",
            "ama category 1": "ama_cat1",
            "category 1 credit": "ama_cat1",
            "category 1": "ama_cat1",
            "ama pra category 2": "ama_cat2",
            "category 2": "ama_cat2",
            # AOA
            "aoa category 1-a": "aoa_1a",
            "aoa 1a": "aoa_1a",
            "aoa category 1-b": "aoa_1b",
            "aoa 1b": "aoa_1b",
            # NP/PA
            "aanp contact": "aanp_contact",
            "aanp": "aanp_contact",
            "aapa category 1": "aapa_cat1",
            "aapa": "aapa_cat1",
            "ancc contact": "ancc_contact",
            "ancc": "ancc_contact",
            "contact hours": "ancc_contact",
            # Other
            "pharmacology": "pharmacology",
            "pharmacotherapeutics": "pharmacology",
            "moc": "moc",
            "maintenance of certification": "moc",
            "self-assessment": "self_assessment",
            "self assessment": "self_assessment",
            "ethics": "ethics",
            "medical ethics": "ethics",
            "pain management": "pain_mgmt",
            "opioid": "pain_mgmt",
            "cne": "cne",
            "continuing nursing": "cne",
        }
        
        extracted_credit_type = ocr_data.get("credit_type", "")
        if extracted_credit_type:
            normalized_type = extracted_credit_type.lower().strip()
            matched = False
            for key, value in credit_type_map.items():
                if key in normalized_type:
                    ocr_data["credit_type_id"] = value
                    matched = True
                    break
            if not matched:
                # Keep the original text for display but don't map
                ocr_data["credit_type_id"] = "ama_cat1"  # Default
                ocr_data["credit_type_original"] = extracted_credit_type
        
        # Determine OCR status based on extraction quality
        fields_extracted = sum(1 for k in ["title", "provider", "credits", "completion_date"] 
                              if ocr_data.get(k) not in [None, "", 0])
        
        if parse_error:
            ocr_status = "failed"
            ocr_error_message = parse_error
        elif fields_extracted >= 3:
            ocr_status = "completed"
        elif fields_extracted >= 1:
            ocr_status = "partial"  # Some data extracted but needs manual review
            ocr_error_message = "Some fields could not be extracted. Please review and edit."
        else:
            ocr_status = "failed"
            ocr_error_message = "Could not extract certificate data. Please enter details manually."
        
        # Update certificate with OCR data
        update_data = {
            "ocr_status": ocr_status,
            "ocr_data": ocr_data,
            "ocr_error": ocr_error_message,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        if ocr_data.get("title"):
            update_data["title"] = str(ocr_data["title"])[:255]  # Limit length
        if ocr_data.get("provider"):
            update_data["provider"] = str(ocr_data["provider"])[:255]
        if ocr_data.get("credits"):
            try:
                credits_val = ocr_data["credits"]
                if isinstance(credits_val, str):
                    # Extract number from string like "1.5 credits"
                    import re
                    num_match = re.search(r'[\d.]+', credits_val)
                    if num_match:
                        credits_val = num_match.group()
                update_data["credits"] = float(credits_val)
            except (ValueError, TypeError):
                pass
        if ocr_data.get("credit_type_id"):
            update_data["credit_type"] = ocr_data["credit_type_id"]
            update_data["credit_types"] = [ocr_data["credit_type_id"]]
        if ocr_data.get("completion_date"):
            # Validate date format
            date_str = str(ocr_data["completion_date"])
            try:
                datetime.strptime(date_str, "%Y-%m-%d")
                update_data["completion_date"] = date_str
            except ValueError:
                # Try to parse other formats
                for fmt in ["%m/%d/%Y", "%d/%m/%Y", "%B %d, %Y", "%b %d, %Y"]:
                    try:
                        parsed = datetime.strptime(date_str, fmt)
                        update_data["completion_date"] = parsed.strftime("%Y-%m-%d")
                        break
                    except ValueError:
                        continue
        if ocr_data.get("certificate_number"):
            update_data["certificate_number"] = str(ocr_data["certificate_number"])[:100]
        if ocr_data.get("subject"):
            update_data["subject"] = str(ocr_data["subject"])[:255]
        
        await db.certificates.update_one(
            {"certificate_id": certificate_id},
            {"$set": update_data}
        )
        
        cert = await db.certificates.find_one({"certificate_id": certificate_id}, {"_id": 0})
        return cert
        
    except Exception as e:
        error_msg = str(e)
        logger.error(f"OCR Error for {certificate_id}: {error_msg}")
        
        # Provide helpful error message to user
        if "rate limit" in error_msg.lower():
            ocr_error_message = "OCR service is busy. Please try again in a moment or enter details manually."
        elif "timeout" in error_msg.lower():
            ocr_error_message = "OCR processing timed out. Please enter details manually."
        elif not ocr_error_message:
            ocr_error_message = "OCR processing failed. Please enter certificate details manually."
        
        await db.certificates.update_one(
            {"certificate_id": certificate_id},
            {"$set": {
                "ocr_status": "failed", 
                "ocr_error": ocr_error_message,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        cert = await db.certificates.find_one({"certificate_id": certificate_id}, {"_id": 0})
        return cert

@api_router.post("/certificates/eeds-import")
async def import_eeds_certificate(request: Request, user: User = Depends(get_current_user)):
    """Import certificate from EEDS QR code data"""
    body = await request.json()
    qr_data = body.get("qr_data", "")
    
    # Parse EEDS QR code data (format varies, this is a basic implementation)
    # EEDS typically encodes certificate info in the QR code
    credit_types = body.get("credit_types", [])
    if not credit_types and body.get("credit_type"):
        credit_types = [body.get("credit_type")]
    
    cert_data = {
        "title": body.get("title", "EEDS Certificate"),
        "provider": body.get("provider", "EEDS"),
        "credits": float(body.get("credits", 1)),
        "credit_types": credit_types,
        "credit_type": credit_types[0] if credit_types else "ama_cat1",
        "completion_date": body.get("completion_date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
        "certificate_number": body.get("certificate_number"),
        "subject": body.get("subject"),
        "eeds_imported": True
    }
    
    cert = Certificate(user_id=user.user_id, **cert_data)
    cert_dict = cert.model_dump()
    cert_dict["created_at"] = cert_dict["created_at"].isoformat()
    cert_dict["updated_at"] = cert_dict["updated_at"].isoformat()
    
    await db.certificates.insert_one(cert_dict)
    cert_dict.pop("_id", None)  # Remove MongoDB's _id to avoid serialization error
    
    # Update requirement progress
    await update_requirement_progress(user.user_id)
    
    return cert_dict

@api_router.post("/certificates/bulk-import")
async def bulk_import_certificates(request: Request, user: User = Depends(get_current_user)):
    """Bulk import certificates from CSV data"""
    body = await request.json()
    certificates_data = body.get("certificates", [])
    
    if not certificates_data:
        raise HTTPException(status_code=400, detail="No certificates provided")
    
    imported = []
    errors = []
    
    for idx, cert_data in enumerate(certificates_data):
        try:
            # Validate required fields
            if not cert_data.get("title"):
                errors.append({"row": idx + 1, "error": "Missing title"})
                continue
            if not cert_data.get("provider"):
                errors.append({"row": idx + 1, "error": "Missing provider"})
                continue
            if not cert_data.get("completion_date"):
                errors.append({"row": idx + 1, "error": "Missing completion date"})
                continue
            
            # Handle credit types
            credit_types = cert_data.get("credit_types", [])
            if isinstance(credit_types, str):
                credit_types = [t.strip() for t in credit_types.split(",") if t.strip()]
            if not credit_types and cert_data.get("credit_type"):
                credit_types = [cert_data.get("credit_type")]
            
            cert = Certificate(
                user_id=user.user_id,
                title=cert_data.get("title", ""),
                provider=cert_data.get("provider", ""),
                credits=float(cert_data.get("credits", 0)),
                credit_types=credit_types,
                credit_type=credit_types[0] if credit_types else "ama_cat1",
                subject=cert_data.get("subject"),
                completion_date=cert_data.get("completion_date"),
                expiration_date=cert_data.get("expiration_date"),
                certificate_number=cert_data.get("certificate_number")
            )
            
            cert_dict = cert.model_dump()
            cert_dict["created_at"] = cert_dict["created_at"].isoformat()
            cert_dict["updated_at"] = cert_dict["updated_at"].isoformat()
            
            await db.certificates.insert_one(cert_dict)
            cert_dict.pop("_id", None)
            imported.append(cert_dict)
            
        except Exception as e:
            errors.append({"row": idx + 1, "error": str(e)})
    
    # Update requirement progress
    await update_requirement_progress(user.user_id)
    
    return {
        "imported_count": len(imported),
        "error_count": len(errors),
        "imported": imported,
        "errors": errors
    }

# ============ REQUIREMENTS ROUTES ============

@api_router.get("/requirements")
async def get_requirements(
    user: User = Depends(get_current_user),
    active_only: bool = True
):
    """Get user's requirements"""
    query = {"user_id": user.user_id}
    if active_only:
        query["is_active"] = True
    
    requirements = await db.requirements.find(query, {"_id": 0}).sort("due_date", 1).to_list(100)
    return requirements

@api_router.post("/requirements")
async def create_requirement(req_data: RequirementCreate, user: User = Depends(get_current_user)):
    """Create a new requirement"""
    req = Requirement(user_id=user.user_id, **req_data.model_dump())
    
    req_dict = req.model_dump()
    req_dict["created_at"] = req_dict["created_at"].isoformat()
    req_dict["updated_at"] = req_dict["updated_at"].isoformat()
    
    await db.requirements.insert_one(req_dict)
    
    # Calculate initial progress
    await update_single_requirement_progress(user.user_id, req.requirement_id)
    
    req = await db.requirements.find_one({"requirement_id": req.requirement_id}, {"_id": 0})
    return req

@api_router.get("/requirements/{requirement_id}")
async def get_requirement(requirement_id: str, user: User = Depends(get_current_user)):
    """Get a specific requirement"""
    req = await db.requirements.find_one(
        {"requirement_id": requirement_id, "user_id": user.user_id},
        {"_id": 0}
    )
    if not req:
        raise HTTPException(status_code=404, detail="Requirement not found")
    return req

@api_router.put("/requirements/{requirement_id}")
async def update_requirement(
    requirement_id: str,
    req_data: RequirementUpdate,
    user: User = Depends(get_current_user)
):
    """Update a requirement"""
    update_dict = {k: v for k, v in req_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.requirements.update_one(
        {"requirement_id": requirement_id, "user_id": user.user_id},
        {"$set": update_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Requirement not found")
    
    req = await db.requirements.find_one({"requirement_id": requirement_id}, {"_id": 0})
    return req

@api_router.delete("/requirements/{requirement_id}")
async def delete_requirement(requirement_id: str, user: User = Depends(get_current_user)):
    """Delete a requirement"""
    result = await db.requirements.delete_one(
        {"requirement_id": requirement_id, "user_id": user.user_id}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Requirement not found")
    
    return {"message": "Requirement deleted"}


@api_router.get("/certificates/filters/options")
async def get_certificate_filter_options(user: User = Depends(get_current_user)):
    """Get unique providers and subjects from user's certificates for autocomplete"""
    # Get unique providers
    providers_pipeline = [
        {"$match": {"user_id": user.user_id}},
        {"$group": {"_id": "$provider"}},
        {"$match": {"_id": {"$ne": None, "$ne": ""}}},
        {"$sort": {"_id": 1}}
    ]
    providers_result = await db.certificates.aggregate(providers_pipeline).to_list(100)
    providers = [p["_id"] for p in providers_result if p["_id"]]
    
    # Get unique subjects
    subjects_pipeline = [
        {"$match": {"user_id": user.user_id}},
        {"$group": {"_id": "$subject"}},
        {"$match": {"_id": {"$ne": None, "$ne": ""}}},
        {"$sort": {"_id": 1}}
    ]
    subjects_result = await db.certificates.aggregate(subjects_pipeline).to_list(100)
    subjects = [s["_id"] for s in subjects_result if s["_id"]]
    
    return {
        "providers": providers,
        "subjects": subjects
    }


async def update_requirement_progress(user_id: str):
    """Update progress for all user requirements"""
    requirements = await db.requirements.find(
        {"user_id": user_id, "is_active": True},
        {"_id": 0}
    ).to_list(100)
    
    for req in requirements:
        await update_single_requirement_progress(user_id, req["requirement_id"])

async def update_single_requirement_progress(user_id: str, requirement_id: str):
    """Update progress for a single requirement with multi-criteria filtering (includes self-reported credits)"""
    req = await db.requirements.find_one({"requirement_id": requirement_id}, {"_id": 0})
    if not req:
        return
    
    # Build base match conditions
    base_match = [{"user_id": user_id}]
    
    # Handle year range filtering
    start_year = req.get("start_year")
    end_year = req.get("end_year")
    
    year_conditions = []
    if start_year:
        year_conditions.append({
            "$expr": {"$gte": [{"$toInt": {"$substr": ["$completion_date", 0, 4]}}, start_year]}
        })
    if end_year:
        year_conditions.append({
            "$expr": {"$lte": [{"$toInt": {"$substr": ["$completion_date", 0, 4]}}, end_year]}
        })
    
    # Handle credit types filtering
    credit_types = req.get("credit_types", [])
    credit_type = req.get("credit_type")
    
    credit_type_condition = None
    if credit_types:
        credit_type_condition = {
            "$or": [
                {"credit_types": {"$in": credit_types}},
                {"credit_type": {"$in": credit_types}}
            ]
        }
    elif credit_type:
        credit_type_condition = {
            "$or": [
                {"credit_types": credit_type},
                {"credit_type": credit_type}
            ]
        }
    
    # Handle provider filtering (case-insensitive partial match) - certificates only
    providers = req.get("providers", [])
    provider_condition = None
    if providers:
        provider_conditions = []
        for provider in providers:
            provider_conditions.append({
                "provider": {"$regex": provider, "$options": "i"}
            })
        provider_condition = {"$or": provider_conditions}
    
    # Handle subject filtering (case-insensitive partial match) - certificates only
    subjects = req.get("subjects", [])
    subject_condition = None
    if subjects:
        subject_conditions = []
        for subject in subjects:
            subject_conditions.append({
                "subject": {"$regex": subject, "$options": "i"}
            })
        subject_condition = {"$or": subject_conditions}
    
    # Build certificate query
    cert_match = base_match.copy()
    cert_match.extend(year_conditions)
    if credit_type_condition:
        cert_match.append(credit_type_condition)
    if provider_condition:
        cert_match.append(provider_condition)
    if subject_condition:
        cert_match.append(subject_condition)
    
    cert_query = {"$and": cert_match} if len(cert_match) > 1 else cert_match[0]
    
    cert_pipeline = [
        {"$match": cert_query},
        {"$group": {"_id": None, "total_credits": {"$sum": "$credits"}, "count": {"$sum": 1}}}
    ]
    
    cert_result = await db.certificates.aggregate(cert_pipeline).to_list(1)
    cert_credits = cert_result[0]["total_credits"] if cert_result else 0
    cert_count = cert_result[0]["count"] if cert_result else 0
    
    # Build self-reported credits query (no provider/subject filtering - just year and credit type)
    self_match = base_match.copy()
    self_match.extend(year_conditions)
    if credit_type_condition:
        self_match.append(credit_type_condition)
    
    self_query = {"$and": self_match} if len(self_match) > 1 else self_match[0]
    
    self_pipeline = [
        {"$match": self_query},
        {"$group": {"_id": None, "total_credits": {"$sum": "$credits"}, "count": {"$sum": 1}}}
    ]
    
    self_result = await db.self_reported_credits.aggregate(self_pipeline).to_list(1)
    self_credits = self_result[0]["total_credits"] if self_result else 0
    self_count = self_result[0]["count"] if self_result else 0
    
    # Total credits
    total_credits = cert_credits + self_credits
    total_count = cert_count + self_count
    
    await db.requirements.update_one(
        {"requirement_id": requirement_id},
        {"$set": {
            "credits_earned": total_credits,
            "matching_certificates": cert_count,
            "matching_self_reported": self_count,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )

# ============ REPORTS ROUTES ============

@api_router.get("/reports/summary")
async def get_report_summary(
    user: User = Depends(get_current_user),
    year: Optional[int] = None
):
    """Get summary report data"""
    current_year = year or datetime.now().year
    
    # Get certificates for the year
    certs = await db.certificates.find(
        {"user_id": user.user_id, "completion_date": {"$regex": f"^{current_year}"}},
        {"_id": 0}
    ).to_list(1000)
    
    # Group by credit type (handle multiple credit_types)
    by_type = {}
    total_credits = 0
    
    for cert in certs:
        credit_types = cert.get("credit_types", [])
        if not credit_types and cert.get("credit_type"):
            credit_types = [cert.get("credit_type")]
        if not credit_types:
            credit_types = ["unknown"]
        
        credits = cert.get("credits", 0)
        total_credits += credits
        
        # Distribute credits to each type (for reporting purposes)
        for credit_type in credit_types:
            if credit_type not in by_type:
                by_type[credit_type] = {"credits": 0, "count": 0}
            by_type[credit_type]["credits"] += credits
            by_type[credit_type]["count"] += 1
    
    # Get requirements progress
    requirements = await db.requirements.find(
        {"user_id": user.user_id, "is_active": True},
        {"_id": 0}
    ).to_list(100)
    
    return {
        "year": current_year,
        "total_certificates": len(certs),
        "total_credits": total_credits,
        "by_credit_type": by_type,
        "requirements": requirements,
        "certificates": certs
    }

@api_router.get("/reports/year-over-year")
async def get_year_over_year_report(
    user: User = Depends(get_current_user),
    start_year: Optional[int] = None,
    end_year: Optional[int] = None
):
    """Get year-over-year comparison data"""
    current_year = datetime.now().year
    end_year = end_year or current_year
    start_year = start_year or (end_year - 4)  # Default to last 5 years
    
    years_data = []
    
    for year in range(start_year, end_year + 1):
        # Get certificates for the year
        certs = await db.certificates.find(
            {"user_id": user.user_id, "completion_date": {"$regex": f"^{year}"}},
            {"_id": 0}
        ).to_list(1000)
        
        # Calculate totals
        total_credits = sum(cert.get("credits", 0) for cert in certs)
        
        # Group by credit type
        by_type = {}
        for cert in certs:
            credit_types = cert.get("credit_types", [])
            if not credit_types and cert.get("credit_type"):
                credit_types = [cert.get("credit_type")]
            if not credit_types:
                credit_types = ["unknown"]
            
            for credit_type in credit_types:
                if credit_type not in by_type:
                    by_type[credit_type] = 0
                by_type[credit_type] += cert.get("credits", 0)
        
        years_data.append({
            "year": year,
            "total_certificates": len(certs),
            "total_credits": total_credits,
            "by_credit_type": by_type
        })
    
    return {
        "start_year": start_year,
        "end_year": end_year,
        "years": years_data
    }

@api_router.get("/reports/export/pdf")
async def export_pdf(
    user: User = Depends(get_current_user),
    year: Optional[int] = None
):
    """Export transcript as PDF"""
    summary = await get_report_summary(user, year)
    
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Header
    c.setFont("Helvetica-Bold", 18)
    c.drawString(1*inch, height - 1*inch, "CME Transcript")
    
    c.setFont("Helvetica", 12)
    c.drawString(1*inch, height - 1.3*inch, f"Name: {user.name}")
    c.drawString(1*inch, height - 1.5*inch, f"Year: {summary['year']}")
    c.drawString(1*inch, height - 1.7*inch, f"Total Credits: {summary['total_credits']}")
    
    # Credits by type
    y = height - 2.2*inch
    c.setFont("Helvetica-Bold", 14)
    c.drawString(1*inch, y, "Credits by Type")
    y -= 0.3*inch
    
    c.setFont("Helvetica", 10)
    for credit_type, data in summary['by_credit_type'].items():
        c.drawString(1.2*inch, y, f"• {credit_type}: {data['credits']} credits ({data['count']} certificates)")
        y -= 0.25*inch
    
    # Certificates list
    y -= 0.3*inch
    c.setFont("Helvetica-Bold", 14)
    c.drawString(1*inch, y, "Certificates")
    y -= 0.3*inch
    
    c.setFont("Helvetica", 9)
    for cert in summary['certificates'][:30]:  # Limit to prevent overflow
        if y < 1*inch:
            c.showPage()
            y = height - 1*inch
            c.setFont("Helvetica", 9)
        
        c.drawString(1.2*inch, y, f"• {cert['title'][:50]} - {cert['credits']} credits ({cert['completion_date']})")
        y -= 0.2*inch
    
    c.save()
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=cme_transcript_{summary['year']}.pdf"}
    )

@api_router.get("/reports/export/excel")
async def export_excel(
    user: User = Depends(get_current_user),
    year: Optional[int] = None
):
    """Export transcript as Excel"""
    summary = await get_report_summary(user, year)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CME Transcript"
    
    # Header
    ws['A1'] = "CME Transcript"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Name: {user.name}"
    ws['A3'] = f"Year: {summary['year']}"
    ws['A4'] = f"Total Credits: {summary['total_credits']}"
    
    # Headers for certificate table
    headers = ["Title", "Provider", "Credits", "Credit Type", "Completion Date", "Certificate #"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Certificate data
    for row, cert in enumerate(summary['certificates'], 7):
        ws.cell(row=row, column=1, value=cert.get('title', ''))
        ws.cell(row=row, column=2, value=cert.get('provider', ''))
        ws.cell(row=row, column=3, value=cert.get('credits', 0))
        ws.cell(row=row, column=4, value=cert.get('credit_type', ''))
        ws.cell(row=row, column=5, value=cert.get('completion_date', ''))
        ws.cell(row=row, column=6, value=cert.get('certificate_number', ''))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=cme_transcript_{summary['year']}.xlsx"}
    )

@api_router.get("/reports/export/html")
async def export_html(
    user: User = Depends(get_current_user),
    year: Optional[int] = None
):
    """Export transcript as printable HTML"""
    summary = await get_report_summary(user, year)
    
    html = f"""
<!DOCTYPE html>
<html>
<head>
    <title>CME Transcript - {user.name}</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; max-width: 800px; margin: 0 auto; padding: 20px; }}
        h1 {{ color: #4F46E5; border-bottom: 2px solid #4F46E5; padding-bottom: 10px; }}
        .header {{ margin-bottom: 30px; }}
        .header p {{ margin: 5px 0; color: #64748b; }}
        .summary {{ background: #f8fafc; padding: 20px; border-radius: 8px; margin-bottom: 30px; }}
        .summary h3 {{ margin-top: 0; color: #334155; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
        th {{ background: #4F46E5; color: white; padding: 12px; text-align: left; }}
        td {{ padding: 10px; border-bottom: 1px solid #e2e8f0; }}
        tr:hover {{ background: #f8fafc; }}
        .credit-type {{ display: inline-block; background: #e0e7ff; color: #3730a3; padding: 2px 8px; border-radius: 4px; font-size: 12px; }}
        @media print {{
            body {{ print-color-adjust: exact; -webkit-print-color-adjust: exact; }}
            .no-print {{ display: none; }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>CME Transcript</h1>
        <p><strong>Name:</strong> {user.name}</p>
        <p><strong>Email:</strong> {user.email}</p>
        <p><strong>Year:</strong> {summary['year']}</p>
        <p><strong>Generated:</strong> {datetime.now().strftime('%B %d, %Y')}</p>
    </div>
    
    <div class="summary">
        <h3>Summary</h3>
        <p><strong>Total Certificates:</strong> {summary['total_certificates']}</p>
        <p><strong>Total Credits:</strong> {summary['total_credits']}</p>
    </div>
    
    <h2>Certificates</h2>
    <table>
        <thead>
            <tr>
                <th>Title</th>
                <th>Provider</th>
                <th>Credits</th>
                <th>Type</th>
                <th>Date</th>
            </tr>
        </thead>
        <tbody>
"""
    
    for cert in summary['certificates']:
        html += f"""
            <tr>
                <td>{cert.get('title', '')}</td>
                <td>{cert.get('provider', '')}</td>
                <td>{cert.get('credits', 0)}</td>
                <td><span class="credit-type">{cert.get('credit_type', '')}</span></td>
                <td>{cert.get('completion_date', '')}</td>
            </tr>
"""
    
    html += """
        </tbody>
    </table>
    
    <p class="no-print" style="margin-top: 30px; text-align: center;">
        <button onclick="window.print()" style="background: #4F46E5; color: white; border: none; padding: 10px 20px; border-radius: 6px; cursor: pointer;">
            Print Transcript
        </button>
    </p>
</body>
</html>
"""
    
    return Response(content=html, media_type="text/html")

@api_router.get("/reports/export/pars")
async def export_pars(
    user: User = Depends(get_current_user),
    year: Optional[int] = None
):
    """Export transcript in ACCME PARS format for annual reporting
    
    PARS (Program and Activity Reporting System) format includes:
    - Activity details required for ACCME compliance reporting
    - Credit types mapped to ACCME standards
    - Provider information with ACCME numbers where available
    """
    summary = await get_report_summary(user, year)
    
    # Create PARS-compliant export workbook
    wb = openpyxl.Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "PARS Summary"
    
    # PARS Header Information
    ws_summary['A1'] = "ACCME PARS Activity Report"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A2'] = f"Reporting Year: {summary['year']}"
    ws_summary['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary['A4'] = f"Physician Name: {user.name}"
    ws_summary['A5'] = f"NPI Number: {user.npi_number or 'Not Provided'}"
    
    # Summary Statistics
    ws_summary['A7'] = "Summary Statistics"
    ws_summary['A7'].font = Font(bold=True)
    ws_summary['A8'] = f"Total CME Activities: {summary['total_certificates']}"
    ws_summary['A9'] = f"Total Credits: {summary['total_credits']}"
    
    # Credits by Type
    ws_summary['A11'] = "Credits by ACCME Category"
    ws_summary['A11'].font = Font(bold=True)
    row = 12
    for ctype, credits in summary['by_credit_type'].items():
        ws_summary.cell(row=row, column=1, value=ctype)
        ws_summary.cell(row=row, column=2, value=credits.get('credits', 0) if isinstance(credits, dict) else credits)
        row += 1
    
    # Activity Details Sheet
    ws_activities = wb.create_sheet("PARS Activities")
    
    # PARS-required columns
    pars_headers = [
        "Activity ID",
        "Activity Title",
        "Activity Type",
        "Provider/Joint Provider",
        "ACCME Provider Number",
        "Credit Type",
        "Credits Claimed",
        "Activity Date",
        "Completion Date",
        "Certificate Number",
        "Subject/Topic",
        "Delivery Format"
    ]
    
    for col, header in enumerate(pars_headers, 1):
        cell = ws_activities.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Activity data
    for row, cert in enumerate(summary['certificates'], 2):
        ws_activities.cell(row=row, column=1, value=cert.get('certificate_id', ''))
        ws_activities.cell(row=row, column=2, value=cert.get('title', ''))
        ws_activities.cell(row=row, column=3, value="Course")  # Default activity type
        ws_activities.cell(row=row, column=4, value=cert.get('provider', ''))
        ws_activities.cell(row=row, column=5, value=cert.get('accme_provider_number', ''))
        
        # Map credit types to ACCME categories
        credit_type = cert.get('credit_type', '')
        accme_credit = credit_type
        if credit_type in ['ama_cat1', 'AMA PRA Category 1']:
            accme_credit = 'AMA PRA Category 1 Credit(s)'
        elif credit_type in ['ama_cat2', 'AMA PRA Category 2']:
            accme_credit = 'AMA PRA Category 2 Credit(s)'
        
        ws_activities.cell(row=row, column=6, value=accme_credit)
        ws_activities.cell(row=row, column=7, value=cert.get('credits', 0))
        ws_activities.cell(row=row, column=8, value=cert.get('completion_date', ''))
        ws_activities.cell(row=row, column=9, value=cert.get('completion_date', ''))
        ws_activities.cell(row=row, column=10, value=cert.get('certificate_number', ''))
        ws_activities.cell(row=row, column=11, value=cert.get('subject', ''))
        ws_activities.cell(row=row, column=12, value='Live' if cert.get('location') else 'Enduring Material')
    
    # Adjust column widths
    for ws in [ws_summary, ws_activities]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=accme_pars_report_{summary['year']}.xlsx"}
    )

# ============ DASHBOARD ROUTES ============

@api_router.get("/dashboard")
async def get_dashboard(user: User = Depends(get_current_user)):
    """Get dashboard data"""
    current_year = datetime.now().year
    
    # Get recent certificates
    recent_certs = await db.certificates.find(
        {"user_id": user.user_id},
        {"_id": 0}
    ).sort("created_at", -1).limit(5).to_list(5)
    
    # Get active requirements with progress
    requirements = await db.requirements.find(
        {"user_id": user.user_id, "is_active": True},
        {"_id": 0}
    ).sort("due_date", 1).to_list(10)
    
    # Get credits by type for current year
    pipeline = [
        {"$match": {"user_id": user.user_id, "completion_date": {"$regex": f"^{current_year}"}}},
        {"$group": {"_id": "$credit_type", "total": {"$sum": "$credits"}, "count": {"$sum": 1}}}
    ]
    credits_by_type = await db.certificates.aggregate(pipeline).to_list(20)
    
    # Get total credits this year
    total_pipeline = [
        {"$match": {"user_id": user.user_id, "completion_date": {"$regex": f"^{current_year}"}}},
        {"$group": {"_id": None, "total": {"$sum": "$credits"}}}
    ]
    total_result = await db.certificates.aggregate(total_pipeline).to_list(1)
    total_credits = total_result[0]["total"] if total_result else 0
    
    # Get upcoming deadlines
    upcoming = [r for r in requirements if r.get("due_date", "") >= datetime.now().strftime("%Y-%m-%d")][:5]
    
    return {
        "user": user.model_dump(),
        "recent_certificates": recent_certs,
        "requirements": requirements,
        "upcoming_deadlines": upcoming,
        "credits_by_type": credits_by_type,
        "total_credits_this_year": total_credits,
        "year": current_year
    }

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
